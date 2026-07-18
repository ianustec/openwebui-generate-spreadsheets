"""
title: Generate Spreadsheets
author: IANUSTEC
author_url: https://ianustec.com
funding_url: https://github.com/ianustec
description: Generate high-quality native Excel (.xlsx) workbooks from a JSON spec - multi-sheet, Excel Tables, live formulas, charts, conditional formatting
requirements: openpyxl, pydantic
required_open_webui_version: 0.4.0
version: 1.0.0
license: MIT
"""

# ============================================================================
# Native XLSX Spreadsheets Engine — open-source OpenWebUI tool
# ----------------------------------------------------------------------------
# Server-side OpenWebUI tool `generate_spreadsheet`. The model emits a structured
# JSON spec ({title, template, theme, sheets:[{name, kind, ...}]}) and this
# engine renders a NATIVE .xlsx with openpyxl: multi-sheet workbooks, typed
# columns, Excel Tables, live formulas, freeze/filter, data validation,
# conditional formatting and native charts.
#
# Formulas are written as native Excel strings (Excel recalculates on open).
# No LibreOffice is required in the OpenWebUI container.
#
# The workbook is saved via the OpenWebUI Files API (with a /cache/files
# fallback) and a clickable download link is emitted in chat.
#
# License: MIT — Copyright (c) IANUSTEC.
# ============================================================================

from __future__ import annotations

import json
import os
import re
import traceback
import unicodedata
import uuid
from copy import deepcopy
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from pydantic import BaseModel, Field

# --- openpyxl (rendering engine) --------------------------------------------
try:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.chart import (  # type: ignore
        AreaChart,
        BarChart,
        LineChart,
        PieChart,
        Reference,
    )
    from openpyxl.chart.label import DataLabelList  # type: ignore
    from openpyxl.chart.series import DataPoint  # type: ignore
    from openpyxl.comments import Comment  # type: ignore
    from openpyxl.formatting.rule import (  # type: ignore
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        FormulaRule,
    )
    from openpyxl.styles import (  # type: ignore
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter  # type: ignore
    from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
    from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore

    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAS_OPENPYXL = False

# --- OpenWebUI Files API ----------------------------------------------------
try:
    from fastapi import UploadFile  # type: ignore
    from starlette.datastructures import Headers  # type: ignore
    from open_webui.routers.files import upload_file_handler  # type: ignore
    from open_webui.models.users import Users  # type: ignore

    _HAS_OWUI_FILES = True
except Exception:
    _HAS_OWUI_FILES = False


# ============================================================================
# Colour / theme helpers
# ============================================================================

_UNSAFE_FORMULA_RE = re.compile(
    r"\b(XLOOKUP|XMATCH|SORT|FILTER|UNIQUE|SEQUENCE|LAMBDA|LET|HSTACK|VSTACK)\s*\(",
    re.IGNORECASE,
)


def _hex(value: str, default: str = "1E2761") -> str:
    if not isinstance(value, str):
        return default
    h = value.strip().lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    if len(h) != 6:
        return default
    try:
        int(h, 16)
        return h.upper()
    except ValueError:
        return default


def _rgb_tuple(hex6: str) -> tuple[int, int, int]:
    h = _hex(hex6)
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=_hex(hex6))


def _font(
    *,
    name: str = "Calibri",
    size: float = 11,
    bold: bool = False,
    color: str = "1F2430",
    italic: bool = False,
) -> Font:
    return Font(
        name=name,
        size=size,
        bold=bold,
        italic=italic,
        color=_hex(color, "1F2430"),
    )


def _thin_border(color: str = "D0D7DE") -> Border:
    side = Side(style="thin", color=_hex(color, "D0D7DE"))
    return Border(left=side, right=side, top=side, bottom=side)


def _slugify(text: str, *, max_len: int = 72) -> str:
    t = unicodedata.normalize("NFKD", text or "")
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = t.lower()
    t = re.sub(r"[^a-z0-9]+", "-", t)
    t = re.sub(r"-{2,}", "-", t).strip("-")
    return (t or "workbook")[:max_len]


def _human_filename(text: str, *, max_len: int = 90) -> str:
    t = (text or "").strip()
    t = re.sub(r'[\\/:*?"<>|\r\n\t]+', " ", t)
    t = re.sub(r"\s{2,}", " ", t).strip(" .-")
    return (t or "Workbook")[:max_len].strip(" .-")


def _deep_merge(base: dict, override: dict) -> dict:
    out = deepcopy(base)
    for k, v in (override or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def _safe_sheet_name(name: str, used: set[str]) -> str:
    raw = re.sub(r'[:\\/?*\[\]]', " ", (name or "Sheet").strip()) or "Sheet"
    raw = raw[:31]
    candidate = raw
    n = 2
    while candidate in used:
        suffix = f" ({n})"
        candidate = (raw[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def _col_letter(idx: int) -> str:
    """1-based column index -> letter."""
    return get_column_letter(idx)


def _quote_sheet(name: str) -> str:
    if re.search(r"[\s'!\-]", name or ""):
        return f"'{name.replace(chr(39), chr(39)+chr(39))}'"
    return name


# ============================================================================
# Templates / themes
# ============================================================================

_NUMBER_FORMATS = {
    "text": "@",
    "number": "#,##0.00",
    "integer": "#,##0",
    "currency": '$#,##0;($#,##0);"-"',
    "currency_eur": '€#,##0;($#,##0);"-"',
    "percent": "0.0%",
    "date": "YYYY-MM-DD",
    "multiple": "0.0x",
}

_TEMPLATE_BLANK: dict = {
    "theme": {
        "font": "Calibri",
        "accent": "#1E2761",
        "header_fg": "#FFFFFF",
        "ink": "#1F2430",
        "muted": "#6B7280",
        "zebra": "#F8FAFC",
        "border": "#D0D7DE",
        "input_fill": "#FFF9C4",
        "input_font": "#0000FF",
        "formula_font": "#1F2430",
        "cross_sheet_font": "#008000",
        "financial_colors": False,
    },
    "defaults": {
        "excel_table": False,
        "auto_filter": True,
        "freeze": "A2",
        "zebra": True,
    },
}

_TEMPLATE_FINANCIAL: dict = {
    "theme": {
        "font": "Calibri",
        "accent": "#1E2761",
        "header_fg": "#FFFFFF",
        "ink": "#1F2430",
        "muted": "#6B7280",
        "zebra": "#F3F6FA",
        "border": "#C9D1D9",
        "input_fill": "#FFFF00",
        "input_font": "#0000FF",
        "formula_font": "#000000",
        "cross_sheet_font": "#008000",
        "financial_colors": True,
    },
    "defaults": {
        "excel_table": True,
        "auto_filter": True,
        "freeze": "A2",
        "zebra": True,
    },
}

_TEMPLATE_REPORT: dict = {
    "theme": {
        "font": "Calibri",
        "accent": "#2E5AAC",
        "header_fg": "#FFFFFF",
        "ink": "#1F2A44",
        "muted": "#64748B",
        "zebra": "#F8FAFC",
        "border": "#D0D7DE",
        "input_fill": "#FFF9C4",
        "input_font": "#0000FF",
        "formula_font": "#1F2430",
        "cross_sheet_font": "#008000",
        "financial_colors": False,
    },
    "defaults": {
        "excel_table": True,
        "auto_filter": True,
        "freeze": "A2",
        "zebra": True,
    },
}

_TEMPLATE_DASHBOARD: dict = {
    "theme": {
        "font": "Calibri",
        "accent": "#0F766E",
        "header_fg": "#FFFFFF",
        "ink": "#123B36",
        "muted": "#64748B",
        "zebra": "#F0FDFA",
        "border": "#99F6E4",
        "input_fill": "#FFF9C4",
        "input_font": "#0000FF",
        "formula_font": "#1F2430",
        "cross_sheet_font": "#008000",
        "financial_colors": False,
    },
    "defaults": {
        "excel_table": True,
        "auto_filter": True,
        "freeze": "A2",
        "zebra": False,
    },
}

TEMPLATES: dict[str, dict] = {
    "blank": _TEMPLATE_BLANK,
    "financial": _TEMPLATE_FINANCIAL,
    "report": _TEMPLATE_REPORT,
    "dashboard": _TEMPLATE_DASHBOARD,
}


def _resolve_template(spec: dict) -> dict:
    name = (spec.get("template") or "financial").lower().strip()
    base = TEMPLATES.get(name, _TEMPLATE_FINANCIAL)
    merged = _deep_merge(base, {k: v for k, v in spec.items() if k != "template"})
    # Coalesce top-level accent into theme
    theme = merged.get("theme") if isinstance(merged.get("theme"), dict) else {}
    if not theme.get("accent"):
        alias = (
            merged.get("accent")
            or merged.get("accent_color")
            or (spec.get("styles") or {}).get("accent")
        )
        if alias:
            theme["accent"] = alias
    merged["theme"] = theme
    merged["template"] = name if name in TEMPLATES else "financial"
    return merged


def _theme(spec: dict) -> dict:
    t = dict(spec.get("theme") or {})
    t.setdefault("font", "Calibri")
    t.setdefault("accent", "#1E2761")
    t.setdefault("header_fg", "#FFFFFF")
    t.setdefault("ink", "#1F2430")
    t.setdefault("muted", "#6B7280")
    t.setdefault("zebra", "#F8FAFC")
    t.setdefault("border", "#D0D7DE")
    t.setdefault("input_fill", "#FFFF00")
    t.setdefault("input_font", "#0000FF")
    t.setdefault("formula_font", "#000000")
    t.setdefault("cross_sheet_font", "#008000")
    t.setdefault("financial_colors", False)
    return t


# ============================================================================
# Kind aliases
# ============================================================================

_KIND_ALIASES = {
    "table": "table",
    "grid": "table",
    "data": "table",
    "matrix": "matrix",
    "pivot_like": "matrix",
    "inputs": "inputs",
    "assumptions": "inputs",
    "input": "inputs",
    "kpi": "kpi_row",
    "kpi_row": "kpi_row",
    "stats": "kpi_row",
    "metrics": "kpi_row",
    "chart": "chart",
    "graph": "chart",
    "notes": "notes",
    "note": "notes",
    "callout": "notes",
    "raw": "raw",
    "cells": "raw",
    "mixed": "mixed",
    "content": "mixed",
}


def _resolve_kind(raw: Optional[str]) -> str:
    if not raw:
        return "mixed"
    return _KIND_ALIASES.get(str(raw).lower().strip(), str(raw).lower().strip())


# ============================================================================
# Cell / formula helpers
# ============================================================================

def _number_format_for(fmt: Optional[str]) -> Optional[str]:
    if not fmt:
        return None
    key = str(fmt).lower().strip()
    if key in _NUMBER_FORMATS:
        return _NUMBER_FORMATS[key]
    # Pass through custom Excel format strings
    return fmt


def _coerce_value(value: Any, fmt: Optional[str] = None) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.startswith("="):
        return value  # formula
    fmt_l = (fmt or "").lower()
    if fmt_l in ("percent",) and isinstance(value, (int, float, str)):
        try:
            v = float(value)
            # If user passed 15 meaning 15%, store 0.15; if already 0.15 keep it
            if abs(v) > 1:
                v = v / 100.0
            return v
        except (TypeError, ValueError):
            return value
    if fmt_l in ("number", "integer", "currency", "currency_eur", "multiple"):
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            cleaned = value.replace(",", "").replace("€", "").replace("$", "").strip()
            try:
                return float(cleaned) if "." in cleaned else int(cleaned)
            except ValueError:
                return value
    if fmt_l == "date" and isinstance(value, str):
        for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, pattern).date()
            except ValueError:
                continue
    return value


def _warn_unsafe_formula(formula: str) -> Optional[str]:
    m = _UNSAFE_FORMULA_RE.search(formula or "")
    if m:
        return m.group(1).upper()
    return None


def _expand_column_formula(
    template: str,
    *,
    row: int,
    columns: list[dict],
    col_index_by_key: dict[str, int],
) -> str:
    """Expand ``=SUM([{q1}]:[{q2}])`` using column keys -> A1 refs for ``row``."""

    def repl(match: re.Match) -> str:
        key = match.group(1)
        idx = col_index_by_key.get(key)
        if idx is None:
            return match.group(0)
        return f"{_col_letter(idx)}{row}"

    out = re.sub(r"\[\{([A-Za-z0-9_]+)\}\]", repl, template)
    # Also support {key} without brackets
    out = re.sub(
        r"\{([A-Za-z0-9_]+)\}",
        lambda m: (
            f"{_col_letter(col_index_by_key[m.group(1)])}{row}"
            if m.group(1) in col_index_by_key
            else m.group(0)
        ),
        out,
    )
    if not out.startswith("="):
        out = "=" + out
    return out


def _apply_cell_style(
    cell,
    *,
    theme: dict,
    bold: bool = False,
    fill_hex: Optional[str] = None,
    font_hex: Optional[str] = None,
    number_format: Optional[str] = None,
    align: Optional[str] = None,
    is_input: bool = False,
    is_formula: bool = False,
    is_cross_sheet: bool = False,
) -> None:
    font_name = theme.get("font") or "Calibri"
    color = font_hex or theme.get("ink") or "1F2430"
    if theme.get("financial_colors"):
        if is_input:
            color = theme.get("input_font") or "0000FF"
        elif is_cross_sheet:
            color = theme.get("cross_sheet_font") or "008000"
        elif is_formula:
            color = theme.get("formula_font") or "000000"
    cell.font = _font(name=font_name, bold=bold, color=color)
    if fill_hex:
        cell.fill = _fill(fill_hex)
    elif is_input and theme.get("financial_colors"):
        cell.fill = _fill(theme.get("input_fill") or "FFFF00")
    cell.border = _thin_border(theme.get("border") or "D0D7DE")
    if number_format:
        cell.number_format = number_format
    if align == "right":
        cell.alignment = Alignment(horizontal="right", vertical="center")
    elif align == "center":
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ============================================================================
# Table renderer
# ============================================================================

def _normalize_columns(raw_cols: list) -> list[dict]:
    cols: list[dict] = []
    for i, c in enumerate(raw_cols or []):
        if isinstance(c, str):
            cols.append({"key": f"c{i}", "header": c, "format": "text"})
        elif isinstance(c, dict):
            key = c.get("key") or c.get("id") or f"c{i}"
            cols.append({
                "key": str(key),
                "header": c.get("header") or c.get("title") or str(key),
                "format": c.get("format") or c.get("type") or "text",
                "width": c.get("width"),
                "formula": c.get("formula"),
                "align": c.get("align"),
                "input": bool(c.get("input") or c.get("is_input")),
            })
    return cols


def _normalize_rows(raw_rows: list, columns: list[dict]) -> list[dict]:
    """Return list of dicts keyed by column key."""
    keys = [c["key"] for c in columns]
    out: list[dict] = []
    for r in raw_rows or []:
        if isinstance(r, dict):
            out.append(r)
        elif isinstance(r, (list, tuple)):
            row: dict = {}
            for i, key in enumerate(keys):
                if i < len(r):
                    row[key] = r[i]
            out.append(row)
    return out


def _write_title_row(ws, title: str, *, theme: dict, start_row: int = 1, cols: int = 1) -> int:
    if not title:
        return start_row
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = _font(
        name=theme.get("font") or "Calibri",
        size=16,
        bold=True,
        color=theme.get("accent") or "1E2761",
    )
    if cols > 1:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=cols,
        )
    return start_row + 1


def _render_table_at(
    ws,
    table_spec: dict,
    *,
    theme: dict,
    defaults: dict,
    start_row: int = 1,
    start_col: int = 1,
    max_rows: int = 5000,
    warnings: list,
) -> dict:
    """Render a table. Returns metadata: {range, header_row, data_start, data_end, cols, name}."""
    columns = _normalize_columns(table_spec.get("columns") or [])
    if not columns:
        return {"end_row": start_row}

    rows = _normalize_rows(table_spec.get("rows") or [], columns)
    if len(rows) > max_rows:
        warnings.append(
            f"Table truncated from {len(rows)} to {max_rows} rows "
            f"(max_rows_per_sheet)."
        )
        rows = rows[:max_rows]

    use_table = table_spec.get("excel_table")
    if use_table is None:
        use_table = defaults.get("excel_table", False)
    auto_filter = table_spec.get("auto_filter")
    if auto_filter is None:
        auto_filter = defaults.get("auto_filter", True)
    zebra = table_spec.get("zebra")
    if zebra is None:
        zebra = defaults.get("zebra", True)
    freeze = table_spec.get("freeze")
    if freeze is None:
        freeze = defaults.get("freeze")
    totals_row = bool(table_spec.get("totals_row"))

    col_index_by_key = {
        c["key"]: start_col + i for i, c in enumerate(columns)
    }
    n_cols = len(columns)

    # Optional title above header
    title = table_spec.get("title") or table_spec.get("caption")
    row_cursor = start_row
    if title:
        row_cursor = _write_title_row(
            ws, str(title), theme=theme, start_row=row_cursor, cols=n_cols,
        )
        row_cursor += 1  # blank spacer under title

    header_row = row_cursor
    accent = theme.get("accent") or "1E2761"
    header_fg = theme.get("header_fg") or "FFFFFF"

    for i, col in enumerate(columns):
        cell = ws.cell(
            row=header_row,
            column=start_col + i,
            value=col["header"],
        )
        cell.font = _font(
            name=theme.get("font") or "Calibri",
            size=11,
            bold=True,
            color=header_fg,
        )
        cell.fill = _fill(accent)
        cell.border = _thin_border(theme.get("border") or "D0D7DE")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = col.get("width")
        if width:
            ws.column_dimensions[_col_letter(start_col + i)].width = float(width)
        else:
            # Heuristic width from header
            ws.column_dimensions[_col_letter(start_col + i)].width = max(
                12, min(40, len(str(col["header"])) + 4)
            )

    ws.row_dimensions[header_row].height = 22

    data_start = header_row + 1
    for r_i, row in enumerate(rows):
        excel_row = data_start + r_i
        for c_i, col in enumerate(columns):
            excel_col = start_col + c_i
            key = col["key"]
            fmt = col.get("format")
            num_fmt = _number_format_for(fmt)
            formula_tmpl = col.get("formula")
            raw_val = row.get(key)

            is_formula = False
            is_cross = False
            if isinstance(raw_val, str) and raw_val.startswith("="):
                value = raw_val
                is_formula = True
            elif formula_tmpl:
                # Column-level formula always wins over static cell values
                value = _expand_column_formula(
                    str(formula_tmpl),
                    row=excel_row,
                    columns=columns,
                    col_index_by_key=col_index_by_key,
                )
                is_formula = True
            else:
                value = _coerce_value(raw_val, fmt)

            if is_formula and isinstance(value, str):
                unsafe = _warn_unsafe_formula(value)
                if unsafe:
                    warnings.append(
                        f"Formula uses {unsafe}; prefer INDEX/MATCH/SUMIFS/IFERROR."
                    )
                if "!" in value:
                    is_cross = True

            cell = ws.cell(row=excel_row, column=excel_col, value=value)
            align = col.get("align")
            if not align and fmt in (
                "number", "integer", "currency", "currency_eur", "percent", "multiple",
            ):
                align = "right"
            is_input = bool(col.get("input") or row.get("_input") or row.get(f"{key}_input"))
            _apply_cell_style(
                cell,
                theme=theme,
                fill_hex=(theme.get("zebra") if zebra and r_i % 2 == 1 else None),
                number_format=num_fmt,
                align=align,
                is_input=is_input,
                is_formula=is_formula,
                is_cross_sheet=is_cross,
            )

    data_end = data_start + len(rows) - 1 if rows else header_row
    end_row = data_end

    # Totals row
    if totals_row and rows:
        end_row = data_end + 1
        totals_spec = table_spec.get("totals") or {}
        for c_i, col in enumerate(columns):
            excel_col = start_col + c_i
            key = col["key"]
            cell = ws.cell(row=end_row, column=excel_col)
            if c_i == 0 and key not in totals_spec:
                cell.value = totals_spec.get("label") or "Total"
            elif key in totals_spec:
                tval = totals_spec[key]
                if isinstance(tval, str) and tval.startswith("="):
                    cell.value = tval
                elif tval in ("sum", "SUM"):
                    letter = _col_letter(excel_col)
                    cell.value = f"=SUM({letter}{data_start}:{letter}{data_end})"
                elif tval in ("avg", "average", "AVERAGE"):
                    letter = _col_letter(excel_col)
                    cell.value = f"=AVERAGE({letter}{data_start}:{letter}{data_end})"
                else:
                    cell.value = tval
            elif col.get("format") in (
                "number", "integer", "currency", "currency_eur",
            ):
                letter = _col_letter(excel_col)
                cell.value = f"=SUM({letter}{data_start}:{letter}{data_end})"
            _apply_cell_style(
                cell,
                theme=theme,
                bold=True,
                fill_hex=theme.get("zebra"),
                number_format=_number_format_for(col.get("format")),
                align="right" if c_i > 0 else "left",
                is_formula=isinstance(cell.value, str) and str(cell.value).startswith("="),
            )

    # Excel Table
    table_name = table_spec.get("name") or table_spec.get("table_name")
    ref = (
        f"{_col_letter(start_col)}{header_row}:"
        f"{_col_letter(start_col + n_cols - 1)}{end_row}"
    )
    if use_table and rows:
        safe_name = re.sub(r"[^A-Za-z0-9_]", "_", str(table_name or f"Table{header_row}"))
        if safe_name[0].isdigit():
            safe_name = "T_" + safe_name
        try:
            tab = Table(displayName=safe_name[:255], ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=bool(zebra),
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            if totals_row:
                tab.totalsRowCount = 1
            ws.add_table(tab)
        except Exception:
            traceback.print_exc()
            # Fall back to auto filter only
            if auto_filter:
                ws.auto_filter.ref = (
                    f"{_col_letter(start_col)}{header_row}:"
                    f"{_col_letter(start_col + n_cols - 1)}{data_end}"
                )
    elif auto_filter and rows:
        ws.auto_filter.ref = (
            f"{_col_letter(start_col)}{header_row}:"
            f"{_col_letter(start_col + n_cols - 1)}{data_end}"
        )

    if freeze:
        try:
            ws.freeze_panes = freeze
        except Exception:
            # freeze like "A2" relative to start — if start_col>1, adjust
            if isinstance(freeze, str) and freeze.upper() == "A2":
                ws.freeze_panes = f"{_col_letter(start_col)}{header_row + 1}"

    # Conditional formatting
    _apply_conditional_formatting(
        ws,
        table_spec.get("conditional_formatting") or table_spec.get("cf") or [],
        header_row=header_row,
        data_start=data_start,
        data_end=data_end,
        start_col=start_col,
        columns=columns,
        col_index_by_key=col_index_by_key,
    )

    # Data validation
    _apply_data_validation(
        ws,
        table_spec.get("validation") or table_spec.get("data_validation") or [],
        data_start=data_start,
        data_end=data_end,
        col_index_by_key=col_index_by_key,
    )

    return {
        "range": ref,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "end_row": end_row + 1,
        "start_col": start_col,
        "n_cols": n_cols,
        "columns": columns,
        "name": table_name,
        "col_index_by_key": col_index_by_key,
    }


# ============================================================================
# Conditional formatting + data validation
# ============================================================================

def _apply_conditional_formatting(
    ws,
    rules: list,
    *,
    header_row: int,
    data_start: int,
    data_end: int,
    start_col: int,
    columns: list,
    col_index_by_key: dict,
) -> None:
    if not rules or data_end < data_start:
        return
    for rule in rules:
        if not isinstance(rule, dict):
            continue
        rtype = (rule.get("type") or rule.get("kind") or "").lower()
        col_key = rule.get("column") or rule.get("col") or rule.get("key")
        if col_key and col_key in col_index_by_key:
            c = col_index_by_key[col_key]
            rng = f"{_col_letter(c)}{data_start}:{_col_letter(c)}{data_end}"
        else:
            rng = rule.get("range") or (
                f"{_col_letter(start_col)}{data_start}:"
                f"{_col_letter(start_col + len(columns) - 1)}{data_end}"
            )
        try:
            if rtype in ("data_bar", "databar"):
                color = _hex(rule.get("color") or "2E5AAC")
                ws.conditional_formatting.add(
                    rng,
                    DataBarRule(
                        start_type="min",
                        end_type="max",
                        color=color,
                        showValue=True,
                        minLength=None,
                        maxLength=None,
                    ),
                )
            elif rtype in ("color_scale", "colorscale", "heatmap"):
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min",
                        start_color=_hex(rule.get("start_color") or "F8696B"),
                        mid_type="percentile",
                        mid_value=50,
                        mid_color=_hex(rule.get("mid_color") or "FFEB84"),
                        end_type="max",
                        end_color=_hex(rule.get("end_color") or "63BE7B"),
                    ),
                )
            elif rtype in ("cell_value", "cellis", "cell_is"):
                op = rule.get("operator") or rule.get("op") or "greaterThan"
                formula = [rule.get("value") or rule.get("formula") or 0]
                fill = rule.get("fill") or rule.get("color")
                font_color = rule.get("font_color")
                kwargs = {}
                if fill:
                    kwargs["fill"] = _fill(fill)
                if font_color:
                    kwargs["font"] = Font(color=_hex(font_color))
                ws.conditional_formatting.add(
                    rng,
                    CellIsRule(operator=op, formula=formula, **kwargs),
                )
            elif rtype in ("formula",):
                f = rule.get("formula") or ""
                if f and not f.startswith("="):
                    f = "=" + f
                fill = rule.get("fill")
                kwargs = {}
                if fill:
                    kwargs["fill"] = _fill(fill)
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f.lstrip("=")], **kwargs),
                )
        except Exception:
            traceback.print_exc()


def _apply_data_validation(
    ws,
    rules: list,
    *,
    data_start: int,
    data_end: int,
    col_index_by_key: dict,
) -> None:
    if not rules or data_end < data_start:
        return
    for rule in rules:
        if not isinstance(rule, dict):
            continue
        col_key = rule.get("column") or rule.get("col") or rule.get("key")
        if not col_key or col_key not in col_index_by_key:
            continue
        c = col_index_by_key[col_key]
        cell_range = f"{_col_letter(c)}{data_start}:{_col_letter(c)}{data_end}"
        vtype = (rule.get("type") or "list").lower()
        try:
            if vtype == "list":
                options = rule.get("options") or rule.get("list") or []
                if isinstance(options, list):
                    formula1 = '"' + ",".join(str(o) for o in options) + '"'
                else:
                    formula1 = str(options)
                dv = DataValidation(
                    type="list",
                    formula1=formula1,
                    allow_blank=bool(rule.get("allow_blank", True)),
                )
            elif vtype in ("whole", "decimal"):
                dv = DataValidation(
                    type=vtype,
                    operator=rule.get("operator") or "between",
                    formula1=str(rule.get("min", rule.get("formula1", 0))),
                    formula2=str(rule.get("max", rule.get("formula2", 100)))
                    if rule.get("max") is not None or rule.get("formula2") is not None
                    else None,
                    allow_blank=True,
                )
            else:
                continue
            dv.error = rule.get("error") or "Invalid value"
            dv.errorTitle = rule.get("error_title") or "Validation"
            dv.add(cell_range)
            ws.add_data_validation(dv)
        except Exception:
            traceback.print_exc()


# ============================================================================
# Other block renderers
# ============================================================================

def _render_inputs(
    ws,
    block: dict,
    *,
    theme: dict,
    start_row: int,
    warnings: list,
) -> int:
    """Assumption / input section: label | value | unit | notes."""
    title = block.get("title") or "Assumptions"
    items = block.get("items") or block.get("inputs") or block.get("rows") or []
    row = _write_title_row(ws, title, theme=theme, start_row=start_row, cols=3)
    row += 1

    legend = block.get("legend") or (
        "Yellow cells with blue text are inputs — edit these. "
        "Black cells are formulas (do not overwrite)."
    )
    ws.cell(row=row, column=1, value=legend).font = _font(
        name=theme.get("font") or "Calibri",
        size=9,
        italic=True,
        color=theme.get("muted") or "6B7280",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    headers = ["Parameter", "Value", "Unit", "Notes"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=1 + i, value=h)
        cell.font = _font(
            name=theme.get("font") or "Calibri",
            bold=True,
            color=theme.get("header_fg") or "FFFFFF",
        )
        cell.fill = _fill(theme.get("accent") or "1E2761")
        cell.border = _thin_border()
    header_row = row
    row += 1

    for item in items:
        if isinstance(item, (list, tuple)):
            label, value, unit, notes = (list(item) + [None, None, None, None])[:4]
            fmt = "number"
            comment = None
        else:
            label = item.get("label") or item.get("name") or item.get("key")
            value = item.get("value")
            unit = item.get("unit")
            notes = item.get("notes") or item.get("note")
            fmt = item.get("format") or "number"
            comment = item.get("comment") or item.get("source")

        ws.cell(row=row, column=1, value=label)
        val_cell = ws.cell(
            row=row,
            column=2,
            value=_coerce_value(value, fmt),
        )
        _apply_cell_style(
            val_cell,
            theme=theme,
            number_format=_number_format_for(fmt),
            is_input=True,
            align="right",
        )
        if comment:
            val_cell.comment = Comment(str(comment), "Assumptions")
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=notes)
        for c in range(1, 5):
            if c != 2:
                _apply_cell_style(ws.cell(row=row, column=c), theme=theme)
        row += 1

    for i, w in enumerate([28, 14, 12, 40]):
        ws.column_dimensions[_col_letter(1 + i)].width = w

    example = block.get("example_row")
    if example:
        row += 1
        ws.cell(row=row, column=1, value="Example (illustrative):").font = _font(
            italic=True, size=9, color=theme.get("muted") or "6B7280",
        )
        row += 1

    freeze = block.get("freeze", f"A{header_row + 1}")
    try:
        ws.freeze_panes = freeze
    except Exception:
        pass

    return row + 1


def _render_kpi_row(
    ws,
    block: dict,
    *,
    theme: dict,
    start_row: int,
) -> int:
    title = block.get("title")
    row = start_row
    if title:
        row = _write_title_row(ws, str(title), theme=theme, start_row=row, cols=4)
        row += 1
    stats = block.get("stats") or block.get("kpis") or block.get("items") or []
    # Layout: each KPI uses 2 columns (label+value stacked via 2 rows)
    col = 1
    for stat in stats:
        if isinstance(stat, dict):
            label = stat.get("label") or ""
            value = stat.get("value")
            change = stat.get("change") or stat.get("delta")
            fmt = stat.get("format")
        else:
            continue
        label_cell = ws.cell(row=row, column=col, value=str(label).upper())
        label_cell.font = _font(
            size=9, bold=True, color=theme.get("muted") or "6B7280",
            name=theme.get("font") or "Calibri",
        )
        val_cell = ws.cell(
            row=row + 1,
            column=col,
            value=_coerce_value(value, fmt),
        )
        val_cell.font = _font(
            size=18,
            bold=True,
            color=theme.get("accent") or "1E2761",
            name=theme.get("font") or "Calibri",
        )
        if fmt:
            val_cell.number_format = _number_format_for(fmt) or val_cell.number_format
        if change:
            ch = ws.cell(row=row + 2, column=col, value=str(change))
            ch.font = _font(size=9, color=theme.get("muted") or "6B7280")
        ws.column_dimensions[_col_letter(col)].width = 18
        # Light card-like fill
        for r in (row, row + 1, row + 2):
            ws.cell(row=r, column=col).fill = _fill(theme.get("zebra") or "F8FAFC")
            ws.cell(row=r, column=col).border = _thin_border(theme.get("border"))
        col += 1
    return row + 4


def _render_notes(
    ws,
    block: dict,
    *,
    theme: dict,
    start_row: int,
) -> int:
    title = block.get("title") or block.get("heading")
    text = block.get("text") or block.get("body") or block.get("note") or ""
    level = (block.get("level") or block.get("kind") or "info").lower()
    colors = {
        "info": ("DBEAFE", "1E40AF"),
        "success": ("DCFCE7", "166534"),
        "warning": ("FEF3C7", "92400E"),
        "danger": ("FEE2E2", "991B1B"),
    }
    bg, fg = colors.get(level, colors["info"])
    row = start_row
    if title:
        cell = ws.cell(row=row, column=1, value=str(title))
        cell.font = _font(bold=True, color=fg, name=theme.get("font") or "Calibri")
        cell.fill = _fill(bg)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    cell = ws.cell(row=row, column=1, value=str(text))
    cell.font = _font(color=fg, name=theme.get("font") or "Calibri")
    cell.fill = _fill(bg)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = max(30, 15 * (1 + str(text).count("\n")))
    return row + 2


def _render_raw(
    ws,
    block: dict,
    *,
    theme: dict,
    start_row: int,
) -> int:
    cells = block.get("cells") or []
    max_row = start_row
    for item in cells:
        if not isinstance(item, dict):
            continue
        # Support A1 or row/col
        addr = item.get("cell") or item.get("addr")
        if addr:
            cell = ws[addr]
            r = cell.row
        else:
            r = int(item.get("row") or start_row)
            c = int(item.get("col") or item.get("column") or 1)
            cell = ws.cell(row=r, column=c)
        value = item.get("value")
        fmt = item.get("format")
        if isinstance(value, str) and value.startswith("="):
            cell.value = value
            is_f = True
        else:
            cell.value = _coerce_value(value, fmt)
            is_f = False
        _apply_cell_style(
            cell,
            theme=theme,
            bold=bool(item.get("bold")),
            fill_hex=item.get("fill"),
            font_hex=item.get("font_color"),
            number_format=_number_format_for(fmt),
            is_input=bool(item.get("input")),
            is_formula=is_f,
            is_cross_sheet="!" in str(value or ""),
        )
        if item.get("comment"):
            cell.comment = Comment(str(item["comment"]), "Note")
        max_row = max(max_row, r)
    for m in block.get("merges") or []:
        try:
            if isinstance(m, str):
                ws.merge_cells(m)
            elif isinstance(m, dict):
                ws.merge_cells(
                    start_row=m["start_row"],
                    start_column=m["start_col"],
                    end_row=m["end_row"],
                    end_column=m["end_col"],
                )
        except Exception:
            traceback.print_exc()
    return max_row + 2


def _render_matrix(
    ws,
    block: dict,
    *,
    theme: dict,
    defaults: dict,
    start_row: int,
    max_rows: int,
    warnings: list,
) -> int:
    """Static pivot-like matrix: headers[] + rows[][] with optional row labels."""
    table = {
        "title": block.get("title"),
        "name": block.get("name") or "Matrix",
        "columns": block.get("columns") or [
            {"key": f"c{i}", "header": h, "format": block.get("format") or "number"}
            for i, h in enumerate(block.get("headers") or [])
        ],
        "rows": block.get("rows") or [],
        "excel_table": block.get("excel_table", False),
        "auto_filter": block.get("auto_filter", True),
        "zebra": True,
    }
    cf = block.get("conditional_formatting")
    if cf is None:
        # Default: color scale on every numeric-looking column except the first
        cf = [
            {"type": "color_scale", "column": c["key"]}
            for c in (table["columns"] or [])[1:]
        ]
    table["conditional_formatting"] = cf
    meta = _render_table_at(
        ws,
        table,
        theme=theme,
        defaults=defaults,
        start_row=start_row,
        max_rows=max_rows,
        warnings=warnings,
    )
    return meta.get("end_row", start_row) + 1


# ============================================================================
# Charts
# ============================================================================

def _render_chart(
    ws,
    block: dict,
    *,
    theme: dict,
    start_row: int,
    table_meta: Optional[dict] = None,
) -> int:
    """Add a native chart. Data from block.labels/values or from table_meta."""
    chart_type = (block.get("chart_type") or block.get("type") or "bar").lower()
    title = block.get("title") or ""

    # Resolve data
    labels = block.get("labels") or block.get("categories")
    values = block.get("values") or block.get("data")
    datasets = block.get("datasets")

    data_start_row = start_row
    if labels and values and not datasets:
        # Write a small data island then chart it
        ws.cell(row=start_row, column=1, value=block.get("category_header") or "Category")
        ws.cell(row=start_row, column=2, value=block.get("value_header") or "Value")
        for i, (lab, val) in enumerate(zip(labels, values)):
            ws.cell(row=start_row + 1 + i, column=1, value=lab)
            ws.cell(row=start_row + 1 + i, column=2, value=_coerce_value(val, "number"))
        n = len(labels)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n)
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + n)
        anchor_row = start_row
        data_end = start_row + n
    elif datasets and labels:
        ws.cell(row=start_row, column=1, value="Category")
        for d_i, ds in enumerate(datasets):
            ws.cell(
                row=start_row,
                column=2 + d_i,
                value=ds.get("label") or ds.get("name") or f"Series {d_i + 1}",
            )
        for i, lab in enumerate(labels):
            ws.cell(row=start_row + 1 + i, column=1, value=lab)
            for d_i, ds in enumerate(datasets):
                series_data = ds.get("data") or ds.get("values") or []
                v = series_data[i] if i < len(series_data) else None
                ws.cell(
                    row=start_row + 1 + i,
                    column=2 + d_i,
                    value=_coerce_value(v, "number"),
                )
        n = len(labels)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n)
        data = Reference(
            ws,
            min_col=2,
            min_row=start_row,
            max_col=1 + len(datasets),
            max_row=start_row + n,
        )
        data_end = start_row + n
        anchor_row = start_row
    elif table_meta and block.get("category_column") and block.get("value_columns"):
        cat_key = block["category_column"]
        val_keys = block["value_columns"]
        if isinstance(val_keys, str):
            val_keys = [val_keys]
        cmap = table_meta.get("col_index_by_key") or {}
        if cat_key not in cmap:
            return start_row
        cat_col = cmap[cat_key]
        val_cols = [cmap[k] for k in val_keys if k in cmap]
        if not val_cols:
            return start_row
        cats = Reference(
            ws,
            min_col=cat_col,
            min_row=table_meta["data_start"],
            max_row=table_meta["data_end"],
        )
        data = Reference(
            ws,
            min_col=min(val_cols),
            max_col=max(val_cols),
            min_row=table_meta["header_row"],
            max_row=table_meta["data_end"],
        )
        data_end = table_meta["data_end"]
        anchor_row = table_meta.get("end_row", start_row)
    else:
        return start_row

    if chart_type in ("pie", "doughnut", "donut"):
        chart = PieChart()
        if chart_type in ("doughnut", "donut"):
            try:
                from openpyxl.chart import DoughnutChart  # type: ignore
                chart = DoughnutChart()
            except Exception:
                chart = PieChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = False
    elif chart_type in ("line",):
        chart = LineChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 10
    elif chart_type in ("area",):
        chart = AreaChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    elif chart_type in ("stacked_bar", "bar_stacked"):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    else:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

    chart.title = title or None
    if chart_type not in ("pie", "doughnut", "donut"):
        if block.get("y_title"):
            chart.y_axis.title = block.get("y_title")
        if block.get("x_title"):
            chart.x_axis.title = block.get("x_title")
    # Accent-ish style
    try:
        chart.style = 10
    except Exception:
        pass

    width = float(block.get("width") or 15)
    height = float(block.get("height") or 8)
    chart.width = width
    chart.height = height

    # Place chart to the right of data or below
    anchor = block.get("anchor") or f"E{anchor_row}"
    ws.add_chart(chart, anchor)
    return max(data_end, anchor_row) + 2


# ============================================================================
# Sheet dispatch
# ============================================================================

def _render_blocks(
    ws,
    blocks: list,
    *,
    theme: dict,
    defaults: dict,
    max_rows: int,
    warnings: list,
) -> None:
    cursor = 1
    last_table_meta = None
    for block in blocks or []:
        if not isinstance(block, dict):
            continue
        kind = _resolve_kind(block.get("kind") or block.get("type") or block.get("layout"))
        try:
            if kind == "table":
                spec = block.get("table") or block
                meta = _render_table_at(
                    ws,
                    spec,
                    theme=theme,
                    defaults=defaults,
                    start_row=cursor,
                    max_rows=max_rows,
                    warnings=warnings,
                )
                last_table_meta = meta
                cursor = meta.get("end_row", cursor) + 2
            elif kind == "matrix":
                cursor = _render_matrix(
                    ws,
                    block,
                    theme=theme,
                    defaults=defaults,
                    start_row=cursor,
                    max_rows=max_rows,
                    warnings=warnings,
                )
            elif kind == "inputs":
                cursor = _render_inputs(
                    ws, block, theme=theme, start_row=cursor, warnings=warnings,
                )
            elif kind == "kpi_row":
                cursor = _render_kpi_row(ws, block, theme=theme, start_row=cursor)
            elif kind == "notes":
                cursor = _render_notes(ws, block, theme=theme, start_row=cursor)
            elif kind == "raw":
                cursor = _render_raw(ws, block, theme=theme, start_row=cursor)
            elif kind == "chart":
                cursor = _render_chart(
                    ws,
                    block,
                    theme=theme,
                    start_row=cursor,
                    table_meta=last_table_meta,
                )
            else:
                # mixed / unknown: try table then notes
                if block.get("columns") or block.get("table"):
                    spec = block.get("table") or block
                    meta = _render_table_at(
                        ws,
                        spec,
                        theme=theme,
                        defaults=defaults,
                        start_row=cursor,
                        max_rows=max_rows,
                        warnings=warnings,
                    )
                    last_table_meta = meta
                    cursor = meta.get("end_row", cursor) + 2
                elif block.get("text") or block.get("body"):
                    cursor = _render_notes(ws, block, theme=theme, start_row=cursor)
        except Exception:
            traceback.print_exc()
            warnings.append(f"Failed to render block kind={kind}")
            continue


def _build_workbook(spec: dict, *, max_rows: int = 5000) -> tuple[bytes, list[str]]:
    if not _HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl is not installed. Add it to the tool requirements."
        )
    warnings: list[str] = []
    theme = _theme(spec)
    defaults = spec.get("defaults") or {}

    wb = Workbook()
    # Remove default sheet; recreate from spec
    default_ws = wb.active
    wb.remove(default_ws)

    sheets = spec.get("sheets") or []
    if not sheets:
        # Convenience: top-level table/columns
        sheets = [{
            "name": "Sheet1",
            "kind": "table",
            "table": {
                "columns": spec.get("columns") or [],
                "rows": spec.get("rows") or [],
            },
        }]

    used_names: set[str] = set()
    accent = theme.get("accent") or "1E2761"

    for sheet_spec in sheets:
        if not isinstance(sheet_spec, dict):
            continue
        name = _safe_sheet_name(
            sheet_spec.get("name") or sheet_spec.get("title") or "Sheet",
            used_names,
        )
        ws = wb.create_sheet(title=name)
        try:
            ws.sheet_properties.tabColor = _hex(accent)
        except Exception:
            pass

        kind = _resolve_kind(
            sheet_spec.get("kind") or sheet_spec.get("type") or sheet_spec.get("layout")
        )

        # Sheet-level title
        cursor = 1
        if sheet_spec.get("title") and kind != "inputs":
            cursor = _write_title_row(
                ws, str(sheet_spec["title"]), theme=theme, start_row=1, cols=6,
            )
            cursor += 1

        if sheet_spec.get("blocks"):
            _render_blocks(
                ws,
                sheet_spec["blocks"],
                theme=theme,
                defaults=defaults,
                max_rows=max_rows,
                warnings=warnings,
            )
        elif kind == "table" or sheet_spec.get("columns") or sheet_spec.get("table"):
            table = sheet_spec.get("table") or {
                "columns": sheet_spec.get("columns"),
                "rows": sheet_spec.get("rows"),
                "name": sheet_spec.get("table_name") or sheet_spec.get("name"),
                "excel_table": sheet_spec.get("excel_table"),
                "auto_filter": sheet_spec.get("auto_filter"),
                "freeze": sheet_spec.get("freeze"),
                "totals_row": sheet_spec.get("totals_row"),
                "totals": sheet_spec.get("totals"),
                "conditional_formatting": sheet_spec.get("conditional_formatting"),
                "validation": sheet_spec.get("validation"),
                "title": None if sheet_spec.get("title") else sheet_spec.get("caption"),
                "zebra": sheet_spec.get("zebra"),
            }
            _render_table_at(
                ws,
                table,
                theme=theme,
                defaults=defaults,
                start_row=cursor,
                max_rows=max_rows,
                warnings=warnings,
            )
        elif kind == "inputs":
            _render_inputs(
                ws, sheet_spec, theme=theme, start_row=cursor, warnings=warnings,
            )
        elif kind == "kpi_row":
            _render_kpi_row(ws, sheet_spec, theme=theme, start_row=cursor)
        elif kind == "chart":
            _render_chart(ws, sheet_spec, theme=theme, start_row=cursor)
        elif kind == "matrix":
            _render_matrix(
                ws,
                sheet_spec,
                theme=theme,
                defaults=defaults,
                start_row=cursor,
                max_rows=max_rows,
                warnings=warnings,
            )
        elif kind == "notes":
            _render_notes(ws, sheet_spec, theme=theme, start_row=cursor)
        elif kind == "raw":
            _render_raw(ws, sheet_spec, theme=theme, start_row=cursor)
        else:
            # empty sheet with note
            ws.cell(row=1, column=1, value=sheet_spec.get("title") or name)

        # Optional sheet-level chart after table
        if sheet_spec.get("chart") and isinstance(sheet_spec["chart"], dict):
            # Find last used row roughly
            last = ws.max_row + 2
            _render_chart(ws, sheet_spec["chart"], theme=theme, start_row=last)

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    # Workbook metadata
    try:
        wb.properties.title = str(spec.get("title") or "")[:200]
        wb.properties.creator = str(spec.get("author") or "Generate Spreadsheets")[:100]
    except Exception:
        pass

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), warnings


# ============================================================================
# Tool reply helpers
# ============================================================================

def _workbook_title(spec: dict) -> str:
    return (
        spec.get("title")
        or spec.get("name")
        or (spec.get("sheets") or [{}])[0].get("name")
        or "workbook"
    )


def _tool_success_reply(fname: str, download_url: str) -> str:
    return (
        "[TOOL_RESULT]\n\n"
        "OUTPUT_FOR_USER — Your **next assistant message** must be **only** "
        "the text between the dashed lines (`---`) below. Copy it exactly "
        "(including the blank line and the `[filename](url)` markdown link). "
        "Do not add summaries, bullet lists, section outlines, or extra "
        "sentences. Do not wrap in code fences. Do not use HTML `<a>`; keep "
        "the markdown link.\n\n"
        "---\n"
        "Here is the Excel workbook:\n\n"
        f"[{fname}]({download_url})\n"
        "---\n\n"
        "If you output anything outside the dashed block, the user may lose "
        "the clickable download link."
    )


# ============================================================================
# Tools class
# ============================================================================

class Tools:
    class Valves(BaseModel):
        default_template: str = Field(
            default="financial",
            description=(
                "Template used when the spec omits `template`. "
                "One of: blank, financial, report, dashboard."
            ),
        )
        xlsx_export_dir: str = Field(
            default="/app/backend/data/cache/files",
            description=(
                "Fallback path for .xlsx export when the OpenWebUI Files API "
                "is unavailable."
            ),
        )
        emit_status: bool = Field(
            default=True,
            description="Emit progress status events during generation.",
        )
        max_rows_per_sheet: int = Field(
            default=5000,
            description="Hard cap on data rows per table (protects token/size).",
        )

    def __init__(self):
        self.valves = self.Valves()
        self.citation = False
        self.tools = [self._tool_descriptor()]

    async def generate_spreadsheet(
        self,
        content: str,
        __event_emitter__=None,
        __request__=None,
        __user__=None,
    ) -> str:
        """Generate a native Excel (.xlsx) workbook from a JSON spec.

        Prefer Excel-2007-era formulas (SUM, SUMIFS, INDEX, MATCH, IFERROR).
        Avoid XLOOKUP / FILTER / UNIQUE / SEQUENCE. Formulas are written as
        native Excel strings; Excel recalculates them when the file is opened
        (cached values may be empty in some previewers until then).

        INPUT — a single JSON object (code fences stripped)::

            {
              "title": "Q3 Budget Pack",
              "author": "Finance",
              "template": "financial",
              "theme": { "accent": "#1E2761", "font": "Calibri" },
              "sheets": [
                {
                  "name": "Assumptions",
                  "kind": "inputs",
                  "items": [
                    {"label": "Revenue growth", "value": 0.08, "format": "percent",
                     "unit": "%", "comment": "Source: board pack"}
                  ]
                },
                {
                  "name": "P&L",
                  "kind": "table",
                  "table": {
                    "name": "PnL",
                    "columns": [
                      {"key": "line", "header": "Line item", "width": 28},
                      {"key": "q1", "header": "Q1", "format": "currency"},
                      {"key": "q2", "header": "Q2", "format": "currency"},
                      {"key": "total", "header": "Total", "format": "currency",
                       "formula": "=SUM([{q1}]:[{q2}])"}
                    ],
                    "rows": [
                      {"line": "Revenue", "q1": 120000, "q2": 135000},
                      {"line": "COGS", "q1": 48000, "q2": 52000}
                    ],
                    "excel_table": true,
                    "freeze": "A2",
                    "totals_row": true,
                    "conditional_formatting": [
                      {"type": "data_bar", "column": "q1"}
                    ]
                  }
                },
                {
                  "name": "Charts",
                  "kind": "chart",
                  "chart_type": "bar",
                  "title": "Revenue by quarter",
                  "labels": ["Q1", "Q2"],
                  "values": [120000, 135000]
                }
              ]
            }

        TEMPLATES: ``blank`` | ``financial`` (blue inputs, yellow fill) |
        ``report`` | ``dashboard``.

        SHEET/BLOCK KINDS: ``table``, ``matrix``, ``inputs``, ``kpi_row``,
        ``chart``, ``notes``, ``raw``, ``mixed`` (blocks[]).

        CHART TYPES: ``bar``, ``stacked_bar``, ``line``, ``area``, ``pie``,
        ``doughnut``.

        COLUMN FORMATS: ``text``, ``number``, ``integer``, ``currency``,
        ``currency_eur``, ``percent``, ``date``, ``multiple``.
        """
        if not _HAS_OPENPYXL:
            return self._error_reply(
                "openpyxl is not installed. OpenWebUI should auto-install it "
                "from the tool requirements."
            )

        await self._emit_status(__event_emitter__, "Parsing spreadsheet spec...", done=False)
        try:
            spec_raw = self._parse_content(content)
        except Exception as exc:
            return self._error_reply(f"Invalid JSON: {exc}")

        if not isinstance(spec_raw, dict):
            return self._error_reply("The `content` parameter must be a JSON object.")

        spec_raw.setdefault("template", self.valves.default_template)
        spec = _resolve_template(spec_raw)

        await self._emit_status(__event_emitter__, "Building Excel workbook...", done=False)
        try:
            xlsx_bytes, warnings = _build_workbook(
                spec, max_rows=int(self.valves.max_rows_per_sheet or 5000),
            )
        except Exception as exc:
            traceback.print_exc()
            return self._error_reply(f"Rendering error: {exc}")

        for w in warnings[:5]:
            await self._emit_status(__event_emitter__, w, done=False)

        await self._emit_status(__event_emitter__, "Saving file...", done=False)
        fname, download_url, save_err = self._save_xlsx(
            xlsx_bytes,
            title=_workbook_title(spec),
            request=__request__,
            user_dict=__user__,
        )
        if not download_url:
            await self._emit_status(__event_emitter__, "Save failed.", done=True)
            return self._error_reply(save_err or "could not save file")

        await self._emit_link(__event_emitter__, fname, download_url)
        await self._emit_status(__event_emitter__, "Workbook ready.", done=True)
        return _tool_success_reply(fname, download_url)

    @staticmethod
    def _parse_content(content) -> Any:
        if isinstance(content, dict):
            return content
        if not isinstance(content, str):
            raise ValueError(f"unsupported content type: {type(content).__name__}")
        text = content.strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```\s*$", "", text)
        text = text.strip()
        return json.loads(text)

    def _save_xlsx(
        self,
        xlsx_bytes: bytes,
        *,
        title: str,
        request=None,
        user_dict=None,
    ) -> tuple[str, Optional[str], Optional[str]]:
        """Persist the XLSX. Returns (display_name, download_url, error).

        Display name is human-readable. Files API uploads under that name.
        /cache/files fallback uses an ASCII slug (static route 500s on spaces).
        """
        display_name = f"{_human_filename(title)}.xlsx"

        if _HAS_OWUI_FILES and request and user_dict:
            try:
                user_model = Users.get_user_by_id(user_dict["id"])
                if user_model:
                    upload = UploadFile(
                        file=BytesIO(xlsx_bytes),
                        filename=display_name,
                        headers=Headers({
                            "content-type": (
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet"
                            ),
                        }),
                    )
                    file_item = upload_file_handler(
                        request=request,
                        file=upload,
                        metadata={},
                        process=False,
                        user=user_model,
                    )
                    if file_item:
                        file_id = getattr(file_item, "id", None)
                        if file_id:
                            return (
                                display_name,
                                f"/api/v1/files/{file_id}/content",
                                None,
                            )
            except Exception:
                traceback.print_exc()

        export_dir = (
            (self.valves.xlsx_export_dir or "").strip()
            or "/app/backend/data/cache/files"
        )
        try:
            os.makedirs(export_dir, mode=0o775, exist_ok=True)
            stored = f"{_slugify(title)}_{uuid.uuid4().hex[:6]}.xlsx"
            filepath = os.path.join(export_dir, stored)
            with open(filepath, "wb") as fh:
                fh.write(xlsx_bytes)
            if os.path.isfile(filepath) and os.path.getsize(filepath) > 0:
                return display_name, f"/cache/files/{stored}", None
        except Exception as exc:
            return display_name, None, str(exc)
        return display_name, None, "could not save file"

    @staticmethod
    def _error_reply(msg: str) -> str:
        return (
            "[TOOL_RESULT — use the text below as your final reply, "
            "verbatim, unchanged. Do NOT include this instruction line.]\n\n"
            f"Could not generate the Excel workbook: {msg}"
        )

    async def _emit_status(self, emitter, description: str, *, done: bool) -> None:
        if not emitter or not self.valves.emit_status:
            return
        try:
            await emitter({
                "type": "status",
                "data": {"description": description, "done": done},
            })
        except Exception:
            pass

    @staticmethod
    async def _emit_link(emitter, fname: str, url: str) -> None:
        if not emitter:
            return
        try:
            await emitter({
                "type": "message",
                "data": {"content": f"\n\n[{fname}]({url})\n"},
            })
        except Exception:
            pass

    @staticmethod
    def _tool_descriptor() -> dict:
        return {
            "type": "function",
            "function": {
                "name": "generate_spreadsheet",
                "description": (
                    "Generate a professional Excel (.xlsx) workbook and save "
                    "it via the OpenWebUI Files API. Multi-sheet workbooks with "
                    "Excel Tables, live formulas, charts, conditional formatting "
                    "and data validation. On success, follow the [TOOL_RESULT] "
                    "OUTPUT_FOR_USER block exactly.\n\n"
                    "Use when the user asks for: Excel, .xlsx, spreadsheet, "
                    "workbook, budget, P&L, forecast, timesheet, multi-sheet "
                    "table. NEVER fabricate an .xlsx yourself — always use this tool.\n\n"
                    "INPUT: JSON object with title/template/theme/sheets[]. "
                    "Sheet kinds: table, inputs, kpi_row, chart, matrix, notes, "
                    "raw, mixed (blocks[]). Templates: blank, financial, report, "
                    "dashboard. Prefer SUM/SUMIFS/INDEX/MATCH/IFERROR formulas; "
                    "avoid XLOOKUP/FILTER/UNIQUE/SEQUENCE."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "content": {
                            "type": "string",
                            "description": (
                                "JSON string (or object) describing the workbook. "
                                "See the tool docstring for the full schema."
                            ),
                        },
                    },
                    "required": ["content"],
                },
            },
        }
